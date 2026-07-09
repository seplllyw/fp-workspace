#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成数据资产板块详细功能清单 Excel"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from spec_rows import ROWS

HEADERS = ["模块", "一级功能", "二级功能", "三级功能", "功能描述", "规则/交互说明", "备注"]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "数据资产功能清单"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0D6E6E", end_color="0D6E6E", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="D0D5DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border

    body_font = Font(name="Arial", size=10)
    alt_fill = PatternFill(start_color="F8FAFB", end_color="F8FAFB", fill_type="solid")
    for r_idx, row in enumerate(ROWS, 2):
        for c_idx, val in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font = body_font
            c.alignment = cell_align
            c.border = border
            if r_idx % 2 == 0:
                c.fill = alt_fill

    widths = [10, 14, 18, 22, 40, 48, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(ROWS)+1}"

    ws2 = wb.create_sheet("统计")
    from collections import Counter
    c1 = Counter(r[1] for r in ROWS)
    ws2.append(["一级功能", "条目数"])
    for k, v in sorted(c1.items(), key=lambda x: -x[1]):
        ws2.append([k, v])
    ws2.append([])
    ws2.append(["合计", len(ROWS)])

    out = "/Users/yaweili/Desktop/数据资产原型/1.xlsx"
    wb.save(out)
    print(f"已生成 {len(ROWS)} 条 → {out}")


if __name__ == "__main__":
    main()
